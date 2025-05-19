import os
import random
import sqlite3
from PIL import Image, ImageDraw, ImageFont
from flask import Flask, redirect, render_template, request, session, jsonify, flash
from flask import send_file
from flask_cors import CORS
from flask_mail import Mail, Message
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = os.urandom(24)
CORS(app)

app.config['MAIL_SERVER'] = 'smtp.mail.ru'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'altynordabot@mail.ru'
app.config['MAIL_PASSWORD'] = 'EE3xi0c9kVXGczYyNNGf'

mail = Mail(app)

min_score = 1
rand_code = 0

def init_db():
    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                first_name TEXT NOT NULL,
                last_name TEXT NOT NULL,
                school_class INTEGER,
                class_letter TEXT,
                email TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                payment_status INTEGER DEFAULT 0,
                score_1 INTEGER DEFAULT 0,
                score_2 INTEGER DEFAULT 0,
                score_3 INTEGER DEFAULT 0,
                test_date TEXT,
                certificate_path TEXT,
                forgot_password TEXT DEFAULT 0,
                place INTEGER DEFAULT 0                
            );
        """)
        conn.commit()

init_db()



def generate_certificate(user_id, user_name, user_score):
    try:
        # Имя и фамилия
        full_name = user_name
        x_offset = 860
        y_offset = 535

        # Загружаем фото
        image = Image.open("static/certificate/certificate.jpg")

        # Создаём объект для рисования
        draw = ImageDraw.Draw(image)

        # Загружаем TTF-шрифт с поддержкой кириллицы
        font_path = "OpenSans-Medium.ttf"  # Положи сюда свой шрифт
        font_size = 50
        font = ImageFont.truetype(font_path, font_size)

        # Позиция — с учётом смещений (x_offset, y_offset)
        x = x_offset
        y = y_offset

        # Нарисовать текст
        draw.text((x, y), full_name, font=font, fill=(0, 0, 0))
        draw.text((800, 750), user_score, font=font, fill=(0, 0, 0))

        # Сохраняем в certificate/1.jpg
        image.save(f"static/certificate/{user_id}/certificate.jpg")

        return f"static/certificate/{user_id}/certificate.jpg"
    except Exception as e:
        print(f"Ошибка генерации: {e}")
        return None



@app.route('/', methods=["POST", "GET"])
def index():
    if request.method == "POST":
        login = request.form["login"].lower()
        password = request.form["password"]

        if login == "admin@admin" and password == "admin":
            session["is_admin"] = True
            return redirect("/teacher")
        else:
            with sqlite3.connect("olympiad.db") as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM users WHERE email = ?", (login,))
                user = cursor.fetchone()

                if user:
                    if user[6] == password:
                        session['user_id'] = user[0]
                        session['email'] = user[5]
                        return redirect("/payment")
                    else:
                        return "<script>alert('Неправильный пароль'); window.location = '/';</script>"
                else:
                    return "<script>alert('Пользователь не найден'); window.location = '/';</script>"
    else:
        if 'user_id' not in session:
            return render_template("main_login.html")
        else:
            return redirect("/home")



@app.route('/registration', methods=["POST", "GET"])
def registration():
    if request.method == "POST":
        name = request.form["first_name"]
        last_name = request.form["last_name"]
        school_class = request.form["class"]
        school_class_letter = request.form["class_letter"]
        email = request.form["email"]
        password = request.form["password"]

        try:
            with sqlite3.connect("olympiad.db") as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO users (first_name, last_name, school_class, class_letter, email, password, payment_status)
                    VALUES (?, ?, ?, ?, ?, ?, 1)
                """, (name, last_name, school_class, school_class_letter, email, password))
                user_id = cursor.lastrowid
                folder_path = f"static/certificate/{user_id}"
                os.makedirs(folder_path, exist_ok=True)
                cursor.execute("""
                    UPDATE users SET certificate_path = ? WHERE id = ?
                """, (folder_path + '/', user_id))
                conn.commit()
            return redirect("/")
        except sqlite3.IntegrityError as e:
            if 'UNIQUE constraint failed: users.email' in str(e):
                return "<script>alert('Ошибка: Email уже используется!'); window.location = '/registration';</script>"
            else:
                return "<script>alert('Ошибка при добавлении пользователя'); window.location = '/registration';</script>"
        except Exception as e:
            return f"Общая ошибка: {e}"


    else:
        return render_template("main_reg.html")


@app.route('/home')
def home():
    if 'user_id' not in session:
        return redirect('/')

    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],))
        user = cursor.fetchone()

    if user[7] == 1:
        return render_template("home.html", user=user)
    else:
        return redirect("/payment")



@app.route('/payment')
def payment():
    if 'user_id' not in session:
        return redirect('/')

    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],))
        user = cursor.fetchone()

    if user[7] == 1:
        return redirect("/home")
    else:
        return render_template("payment.html", user=user)


@app.route("/profile/<id>")
def profile(id):
    if 'user_id' not in session:
        return redirect('/')

    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],))
        user = cursor.fetchone()

        if not user:
            return "Пользователь не найден", 404

        total_score = user[8] + user[9] + user[10]
        max_score = 75
        status = "Не пройден"
        cert_path = None

        #if user[8] >= min_score and user[9] >= min_score and user[10] >= min_score:
        if user[14] >= 1:
            status = "Пройден"
            cert_path = f"static/certificate/{user[0]}/certificate.jpg"
            if not os.path.exists(cert_path):
                generated_path = generate_certificate(user[0], f"{user[1]} {user[2]}", str(user[14]))
                if generated_path:
                    cursor.execute("UPDATE users SET certificate_path=? WHERE id=?",
                                 (f"static/certificate/{user[0]}/", user[0]),)
                    conn.commit()
                    cert_path = generated_path

        user_data = {
            'id': user[0],
            'full_name': f"{user[1]} {user[2]}",
            'class_info': f"{user[3]}{user[4]}",
            'email': user[5],
            'payment_status': "Оплачено" if user[7] else "Не оплачено",
            'scores': {
                'test1': user[8],
                'test2': user[9],
                'test3': user[10],
                'total': total_score,
                'max': max_score,
                'status': status
            },
            'certificate_path': f"/static/certificate/{id}/certificate.jpg"
        }

    return render_template('profile.html', user=user_data, min_score=min_score)

@app.route('/verify/<cert_id>')
def verify_certificate(cert_id):
    cert_path = f"static/certificate/{cert_id}/certificate.jpg"
    if os.path.exists(cert_path):
        with sqlite3.connect("olympiad.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT first_name, last_name FROM users WHERE id=?", (cert_id,))
            user = cursor.fetchone()

        if user:
            return render_template("verify.html",
                                   valid=True,
                                   name=f"{user[0]} {user[1]}",
                                   cert_id=cert_id)

    return render_template("verify.html", valid=False)


@app.route('/test_cert/<user_id>')
def test_cert(user_id):
    if not os.environ.get('DEBUG'):  # Доступно только в режиме разработки
        return "Not allowed", 403

    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id=?", (user_id,))
        user = cursor.fetchone()

    if user:
        generate_certificate(user_id, f"{user[1]} {user[2]}", user[14], fake_mode=True)
        return redirect(f"/static/certificates/{user_id}/certificate.jpg")

    return "User not found", 404

@app.route('/print_certificate/<user_id>')
def print_certificate(user_id):
        cert_path = f"static/certificate/{user_id}/certificate.jpg"
        if os.path.exists(cert_path):
            return send_file(cert_path, mimetype='image/png')
        else:
            return "Сертификат не найден", 404


@app.route('/teacher', methods=["POST", "GET"])
def teacher_cabinet():
    if not session.get("is_admin"):
        return redirect("/")
    try:
        with sqlite3.connect("olympiad.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users")
            rows = cursor.fetchall()
    except Exception as e:
        print(f"Ошибка базы данных: {e}")
    return render_template('Tcabinet.html', rows=rows)



@app.route('/teacher/<id>')
def teacher(id):
    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (id,))
        user = cursor.fetchone()


    return render_template('detailed.html', user=user)


@app.route('/teacher/<email>')
def teacher_email(email):
    email = email.strip().lower()
    print(f"Поиск по email: {email}")
    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (email,))
        user = cursor.fetchone()
        print(user)
    return render_template('detailed.html', user=user)


@app.route('/teacher/<int:id>/update', methods=['POST'])
def update_teacher_permission(id):
    data = request.get_json()
    permission_value = data.get('payment_status')

    if permission_value not in [0, 1]:
        return jsonify({'error': 'Неверное значение'}), 400

    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET payment_status = ? WHERE id = ?", (permission_value, id))
        conn.commit()

    return '', 200


@app.route("/update_place/<user_id>", methods=["POST"])
def update_place(user_id):
    data_place = request.get_json()
    place = data_place.get("place")

    if place is not None:
        try:
            with sqlite3.connect("olympiad.db") as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE users
                    SET place = ?
                    WHERE id = ?
                """, (place, user_id))
                conn.commit()
            return "", 204
        except Exception as e:
            print(f"Ошибка при обновлении балла: {e}")
            return "Ошибка при сохранении", 500
    else:
        return "Нет данных о балле", 400


@app.route("/test/1")
def test_1():
    if 'user_id' not in session:
        return redirect('/')

    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],))
        user = cursor.fetchone()

    if user[8] >= min_score:
        return "<script>alert('Вы уже сдали данный тест'); window.location = '/home';</script>"
    else:
        return render_template("test_1.html", user=user, min_score=min_score)


@app.route("/test/2")
def test_2():
    if 'user_id' not in session:
        return redirect('/')

    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],))
        user = cursor.fetchone()

    if user[9] >= min_score:
        return "<script>alert('Вы уже сдали данный тест'); window.location = '/home';</script>"
    else:
        return render_template("test_2.html", user=user, min_score=min_score)


@app.route("/test/3")
def test_3():
    if 'user_id' not in session:
        return redirect('/')

    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],))
        user = cursor.fetchone()

    if user[10] >= min_score:
        return "<script>alert('Вы уже сдали данный тест'); window.location = '/home';</script>"
    else:
        return render_template("test_3.html", user=user, min_score=min_score)


@app.route("/receive_score_1", methods=["POST"])
def receive_score_1():
    if 'user_id' not in session:
        return "Unauthorized", 401

    data = request.get_json()
    score = data.get("score")

    if score is not None:
        try:
            with sqlite3.connect("olympiad.db") as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE users
                    SET score_1 = ?
                    WHERE id = ?
                """, (score, session['user_id']))
                conn.commit()
            return "", 204
        except Exception as e:
            print(f"Ошибка при обновлении балла: {e}")
            return "Ошибка при сохранении", 500
    else:
        return "Нет данных о балле", 400

@app.route("/receive_score_2", methods=["POST"])
def receive_score_2():
    if 'user_id' not in session:
        return "Unauthorized", 401

    data = request.get_json()
    score = data.get("score")

    if score is not None:
        try:
            with sqlite3.connect("olympiad.db") as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE users
                    SET score_2 = ?
                    WHERE id = ?
                """, (score, session['user_id']))
                conn.commit()
            return "", 204
        except Exception as e:
            print(f"Ошибка при обновлении балла: {e}")
            return "Ошибка при сохранении", 500
    else:
        return "Нет данных о балле", 400


@app.route("/receive_score_3", methods=["POST"])
def receive_score_3():
    if 'user_id' not in session:
        return "Unauthorized", 401

    data = request.get_json()
    score = data.get("score")

    if score is not None:
        try:
            with sqlite3.connect("olympiad.db") as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE users
                    SET score_3 = ?
                    WHERE id = ?
                """, (score, session['user_id']))
                conn.commit()
            return "", 204
        except Exception as e:
            print(f"Ошибка при обновлении балла: {e}")
            return "Ошибка при сохранении", 500
    else:
        return "Нет данных о балле", 400

@app.route("/update-pay")
def pay():
    id = session['user_id']
    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE users
            SET payment_status = ?
            WHERE id = ?
        """, (1, id))

        conn.commit()

    return redirect("/home")




@app.route('/upload', methods=['POST'])
def upload_file():
    if 'user_id' not in session:
        return redirect('/')

    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],))
        user = cursor.fetchone()



@app.route("/forgot-password/email", methods=["POST", "GET"])
def forgot_password_email():
    if request.method == "POST":
        email = request.form["email"].lower()

        rand_code = random.randint(111111, 999999)

        with sqlite3.connect("olympiad.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE email = ?", (email,))
            user = cursor.fetchone()

        if user:
            with sqlite3.connect("olympiad.db") as conn:
                cursor = conn.cursor()

                cursor.execute("""
                    UPDATE users
                    SET forgot_password = ?
                    WHERE email = ?
                """, (rand_code, email))

                conn.commit()

            msg = Message("Код для восстанавления аккаунта", sender=app.config['MAIL_USERNAME'], recipients=[email])
            msg.body = f"""Здравствуйте!
    
    Вы запросили восстановление доступа к вашему аккаунту.
    Ваш код для восстановления:
    
    🔑 { rand_code }
    
    Пожалуйста, введите этот код на сайте, чтобы продолжить процесс восстановления.
    
    Если вы не запрашивали сброс пароля, просто проигнорируйте это сообщение.
    
    С уважением,
    Команда поддержки"""
            mail.send(msg)
            return redirect(f"/forgot-password/code/{user[5]}")
        else:
            return "<script>alert('Пользователь не найден'); window.location = '/';</script>"
    else:
        return render_template("forgot_password.html")


@app.route("/forgot-password/code/<email>", methods=["POST", "GET"])
def forgot_password_code(email):
    if request.method == "POST":
        code = request.form["code"]

        with sqlite3.connect("olympiad.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE email = ?", (email,))
            user = cursor.fetchone()

        if code == user[13]:
            return redirect(f"/forgot-password/password/{user[5]}")
        else:
            return "<script>alert('Код неправильный');"
    else:
        return render_template("code.html")



@app.route("/forgot-password/password/<email>", methods=["POST", "GET"])
def forgot_password_password(email):
    if request.method == "POST":
        password = request.form["new_password"]

        with sqlite3.connect("olympiad.db") as conn:
            cursor = conn.cursor()

            cursor.execute("""
                UPDATE users
                SET password = ?
                WHERE email = ?
            """, (password, email))

            conn.commit()

        with sqlite3.connect("olympiad.db") as conn:
            cursor = conn.cursor()

            cursor.execute("""
                UPDATE users
                SET forgot_password = ?
                WHERE email = ?
            """, (0, email))

            conn.commit()


        return redirect("/")
    else:
        return render_template("password.html")


@app.route("/teacher/xlsx")
def download_xlsx():
    file_path = "Участники.xlsx"

    if os.path.exists(file_path):
        os.remove(file_path)

    book = Workbook()
    sheet = book.active

    sheet.cell(row=1, column=1, value="№")
    sheet.cell(row=1, column=2, value="Аты-жөні")
    sheet.cell(row=1, column=3, value="Сыныбы")
    sheet.cell(row=1, column=4, value="Жалпы ұпайы")
    sheet.cell(row=1, column=5, value="Орын")

    with sqlite3.connect("olympiad.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users")
        users = cursor.fetchall()

    for i, user in enumerate(users, start=2):
        sheet.cell(row=i, column=1, value=user[0])
        sheet.cell(row=i, column=2, value=f"{user[1]} {user[2]}")
        sheet.cell(row=i, column=3, value=f'{user[3]} "{user[4]}"')
        sheet.cell(row=i, column=4, value=user[8] + user[9] + user[10])
        sheet.cell(row=i, column=5, value=user[14] if user[14] != 0 else " ")

    book.save(file_path)

    return send_file(file_path, as_attachment=True)




@app.route('/logout')
def logout():
    session.clear()
    return redirect("/")

@app.route("/teacher/delete/db/init/new/db")
def delete_db():
    if not session.get("is_admin"):
        return redirect("/")
    os.remove("olympiad.db")
    init_db()
    return redirect("/")



@app.route("/teacher/download/db")
def download_db():
    if not session.get("is_admin"):
        return redirect("/")
    return send_file('olympiad.db', as_attachment=True)


if __name__ == "__main__":
    app.run(port=5005)